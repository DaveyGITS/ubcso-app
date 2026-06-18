from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models
from functools import wraps
from django.db.models import Q
from django_ratelimit.decorators import ratelimit
from core.audit import log_action, AuditActions

def home_view(request):
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    from organizations.models import Organization
    from django.utils import timezone
    org_count = Organization.objects.filter(is_active=True, is_cso=False).count()
    return render(request, 'core/home.html', {
        'org_count': org_count,
        'year': timezone.now().year,
    })


@login_required
def dashboard_view(request):
    from memberships.models import Membership, MembershipRequest
    from announcements.models import Notification
    from elections.models import Election, ElectionVoter
    from core.enforcement import (
        expire_cso_admin_for_user,
        expire_all_co_chairmen,
        close_overdue_elections,
    )

    # Enforce all timed features on every dashboard load
    expire_cso_admin_for_user(request.user)
    expire_all_co_chairmen()
    close_overdue_elections()

    active_memberships = Membership.objects.filter(
        user=request.user,
        status='active',
        organization__is_active=True,
    ).select_related('organization', 'custom_role')

    orgs_count = active_memberships.count()

    pending_requests = MembershipRequest.objects.filter(
        user=request.user,
        status='pending'
    ).count()

    unread_notifications = Notification.objects.filter(
        recipients__user=request.user
    ).exclude(
        reads__user=request.user
    ).count()

    # Open elections the user can vote in
    open_voter_elections = Election.objects.filter(
        voters__user=request.user,
        status='open',
    ).select_related('organization').order_by('-created_at')

    # Count user's board posts
    from padlet.models import Post
    board_posts_count = Post.objects.filter(author=request.user).count()

    return render(request, 'core/dashboard.html', {
        'orgs_count': orgs_count,
        'pending_count': pending_requests,
        'notifications_count': unread_notifications,
        'board_posts_count': board_posts_count,
        'active_memberships': active_memberships,
        'open_voter_elections': open_voter_elections,
    })


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        # Enforce CSO admin expiry on every admin request
        from core.enforcement import expire_cso_admin_for_user
        expire_cso_admin_for_user(request.user)
        if not request.user.is_cso_admin and not request.user.is_cso_president:
            messages.error(request, 'You do not have permission to access the admin panel.')
            return redirect('core:dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@admin_required
def admin_panel_view(request):
    from organizations.models import Organization, OrganizationRegistration, AccreditationApplication
    from memberships.models import MembershipRequest
    from accounts.models import User
    from reports.models import AccomplishmentReport
    from accounts.models import ProfileCorrectionRequest

    total_students = User.objects.filter(is_active=True).count()
    total_orgs = Organization.objects.filter(is_active=True, is_cso=False).count()
    pending_registration_claims = OrganizationRegistration.objects.filter(status='pending').count()
    pending_accreditation = AccreditationApplication.objects.filter(status__in=['pending', 'under_review']).count()
    pending_membership_requests = MembershipRequest.objects.filter(status='pending').count()
    pending_reports_count = AccomplishmentReport.objects.filter(status='pending').count()
    pending_correction_requests_count = ProfileCorrectionRequest.objects.filter(status='pending').count()

    from accounts.models import FacultyRegistrationRequest
    pending_faculty_requests = FacultyRegistrationRequest.objects.filter(status='pending_approval').count()

    # School year cycle state for the Close Renewal button
    from core.models import SystemSetting
    school_year_transitioned = SystemSetting.objects.filter(key='school_year_transitioned').exists()
    renewal_period_closed = SystemSetting.objects.filter(key='renewal_period_closed').exists()
    renewal_due_count = Organization.objects.filter(status='renewal_due', is_cso=False).count()

    return render(request, 'core/admin/panel.html', {
        'total_students': total_students,
        'total_orgs': total_orgs,
        'pending_registration_claims': pending_registration_claims,
        'pending_accreditation': pending_accreditation,
        'pending_membership_requests': pending_membership_requests,
        'pending_reports_count': pending_reports_count,
        'pending_correction_requests_count': pending_correction_requests_count,
        'pending_faculty_requests': pending_faculty_requests,
        'school_year_transitioned': school_year_transitioned,
        'renewal_period_closed': renewal_period_closed,
        'renewal_due_count': renewal_due_count,
    })


@admin_required
def admin_org_requests_view(request):
    from organizations.models import OrganizationRequest
    requests = OrganizationRequest.objects.all().select_related(
        'requester', 'proposed_chairman'
    ).order_by('-created_at')
    return render(request, 'core/admin/org_requests.html', {'requests': requests})


@admin_required
@ratelimit(key='user', rate='20/h', method='POST', block=True)
def admin_approve_org_request_view(request, request_id):
    from organizations.models import OrganizationRequest, Organization
    from memberships.models import Membership

    org_request = get_object_or_404(OrganizationRequest, id=request_id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            org = Organization.objects.create(
                name=org_request.organization_name,
                description=org_request.description,
                status='active',
            )
            chairman = org_request.proposed_chairman or org_request.requester
            Membership.objects.create(
                user=chairman,
                organization=org,
                role='chairman',
                has_chairman_privileges=True,
                status='active',
            )
            org_request.status = 'approved'
            org_request.created_organization = org
            org_request.reviewed_by = request.user
            from django.utils import timezone
            org_request.reviewed_at = timezone.now()
            org_request.save()
            
            # Audit log
            log_action(
                actor=request.user,
                action=AuditActions.ORG_APPROVED,
                target=org,
                details=f'Approved organization: {org.name}. Chairman: {chairman.get_full_name()}',
                request=request
            )
            
            from announcements.utils import send_notification
            from django.urls import reverse
            recipients = list({org_request.requester, chairman})
            send_notification(
                title=f'Organization approved — {org.name}',
                message=f'Your organization request for "{org.name}" has been approved. {chairman.get_full_name()} has been assigned as chairman.',
                recipients=recipients,
                sender=request.user,
                organization=org,
                is_priority=True,
                link_url=reverse('organizations:org_profile', kwargs={'org_id': org.id}),
                notification_type='organization',
            )
            messages.success(request, f'Organization "{org.name}" approved and created successfully!')

        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason', '').strip()
            org_request.status = 'rejected'
            org_request.rejection_reason = rejection_reason
            org_request.reviewed_by = request.user
            from django.utils import timezone
            org_request.reviewed_at = timezone.now()
            org_request.save()
            
            # Audit log
            log_action(
                actor=request.user,
                action=AuditActions.ORG_REJECTED,
                target=org_request.requester,
                details=f'Rejected organization request: {org_request.organization_name}. Reason: {rejection_reason}',
                request=request
            )
            
            from announcements.utils import send_notification
            from django.urls import reverse
            msg = f'Your organization request for "{org_request.organization_name}" was not approved.'
            if rejection_reason:
                msg += f' Reason: {rejection_reason}'
            send_notification(
                title=f'Organization request rejected — {org_request.organization_name}',
                message=msg,
                recipients=[org_request.requester],
                sender=request.user,
                link_url=reverse('core:dashboard'),
                notification_type='organization',
            )
            messages.success(request, f'Organization request rejected.')

    return redirect('core:admin_org_requests')


@admin_required
def admin_membership_requests_view(request):
    from memberships.models import MembershipRequest
    requests = MembershipRequest.objects.filter(
        type='request'
    ).select_related(
        'user', 'organization'
    ).order_by('-requested_at')
    return render(request, 'core/admin/membership_requests.html', {'requests': requests})


@admin_required
def admin_faculty_requests_view(request):
    """CSO Admin: list, approve, and reject pending faculty registration requests."""
    from accounts.models import FacultyRegistrationRequest, User
    from django.utils import timezone

    if request.method == 'POST':
        req_id = request.POST.get('request_id')
        action = request.POST.get('action')
        faculty_req = get_object_or_404(FacultyRegistrationRequest, id=req_id, status='pending_approval')

        if action == 'approve':
            # Create the faculty User account
            import secrets as _secrets
            user = User.objects.create_user(
                email=faculty_req.email,
                password=None,
                first_name=faculty_req.first_name,
                middle_initial=faculty_req.middle_initial,
                last_name=faculty_req.last_name,
                employee_id=faculty_req.employee_id,
                department=faculty_req.department,
                is_faculty=True,
                is_active=True,
                is_email_verified=True,
            )
            # Set the password from the stored hash directly
            user.password = faculty_req.password_hash
            user.save(update_fields=['password'])

            faculty_req.status = 'approved'
            faculty_req.reviewed_by = request.user
            faculty_req.reviewed_at = timezone.now()
            faculty_req.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])

            # Notify faculty member
            try:
                from django.core.mail import send_mail
                send_mail(
                    subject='UB-CSO Portal — Faculty Account Approved',
                    message=(
                        f'Dear {faculty_req.get_full_name()},\n\n'
                        'Your faculty registration request has been approved. '
                        'You can now log in at the UB-CSO Portal using your email and password.\n\n'
                        'Welcome!'
                    ),
                    from_email='noreply@universityofbohol.edu.ph',
                    recipient_list=[faculty_req.email],
                    fail_silently=True,
                )
            except Exception:
                pass

            log_action(
                actor=request.user,
                action=AuditActions.USER_REGISTERED,
                target=user,
                details=f'Approved faculty registration for {user.get_full_name()} ({user.employee_id})',
                request=request,
            )
            messages.success(request, f'Faculty account for {faculty_req.get_full_name()} approved.')

        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '').strip()
            faculty_req.status = 'rejected'
            faculty_req.reviewed_by = request.user
            faculty_req.reviewed_at = timezone.now()
            faculty_req.rejection_reason = reason
            faculty_req.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'rejection_reason'])

            try:
                from django.core.mail import send_mail
                msg = f'Dear {faculty_req.get_full_name()},\n\nYour faculty registration request has been rejected.'
                if reason:
                    msg += f'\n\nReason: {reason}'
                send_mail(
                    subject='UB-CSO Portal — Faculty Registration Update',
                    message=msg,
                    from_email='noreply@universityofbohol.edu.ph',
                    recipient_list=[faculty_req.email],
                    fail_silently=True,
                )
            except Exception:
                pass

            messages.success(request, f'Faculty request from {faculty_req.get_full_name()} rejected.')

        return redirect('core:admin_faculty_requests')

    requests_qs = FacultyRegistrationRequest.objects.order_by(
        models.Case(models.When(status='pending_approval', then=0), default=1, output_field=models.IntegerField()),
        '-created_at',
    )
    pending_count = requests_qs.filter(status='pending_approval').count()

    return render(request, 'core/admin/faculty_requests.html', {
        'faculty_requests': requests_qs,
        'pending_count': pending_count,
    })


@admin_required
def admin_students_view(request):
    from accounts.models import User
    from memberships.models import Membership
    from django.db.models import Exists, OuterRef

    query = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '').strip()

    students = User.objects.filter(is_active=True).select_related('course').annotate(
        is_adviser=Exists(
            Membership.objects.filter(
                user=OuterRef('pk'),
                role='adviser',
                status='active',
            )
        )
    ).order_by('last_name')

    if query:
        students = students.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(student_id__icontains=query)
        )

    if role_filter == 'faculty':
        students = students.filter(is_faculty=True)

    return render(request, 'core/admin/students.html', {
        'students': students,
        'query': query,
        'role_filter': role_filter,
    })


@admin_required
@ratelimit(key='user', rate='20/h', method='POST', block=True)
def admin_deactivate_student_view(request, user_id):
    from accounts.models import User
    student = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        from memberships.models import Membership, MembershipRequest, LeaveRequest
        from elections.models import ElectionVoter

        # Mark all active memberships as removed
        Membership.objects.filter(user=student, status='active').update(status='removed')

        # Cancel all pending membership requests and invites
        MembershipRequest.objects.filter(user=student, status='pending').update(status='rejected')

        # Remove pending leave requests
        LeaveRequest.objects.filter(user=student, status='pending').delete()

        # Remove from open/draft election voter pools
        ElectionVoter.objects.filter(
            user=student,
            election__status__in=['draft', 'open'],
        ).delete()

        # Deactivate the account
        student.is_active = False
        student.save()

        # Audit log
        log_action(
            actor=request.user,
            action=AuditActions.USER_DEACTIVATED,
            target=student,
            details=f'Deactivated student account: {student.get_full_name()} ({student.student_id})',
            request=request
        )

        messages.success(request, f'{student.get_full_name()} has been deactivated.')
    return redirect('core:admin_students')


@admin_required
def admin_orgs_view(request):
    from organizations.models import Organization
    from django.db.models import Count, Q

    group_filter = request.GET.get('group', '')

    orgs = Organization.objects.filter(is_cso=False).annotate(
        active_member_count=Count(
            'memberships',
            filter=Q(memberships__status='active', memberships__user__is_active=True)
        )
    )

    ACCREDITED_STATUSES = ['active', 'renewal_due', 'lapsed']
    ACCREDITED_CATEGORIES = ['student', 'ub_chapter', 'institutional']
    PROBATIONARY_STATUSES = ['probationary']
    PROBATIONARY_CATEGORIES = ['student', 'ub_chapter']

    if group_filter == 'accredited':
        orgs = orgs.filter(status__in=ACCREDITED_STATUSES, category__in=ACCREDITED_CATEGORIES)
    elif group_filter == 'probationary':
        orgs = orgs.filter(status__in=PROBATIONARY_STATUSES, category__in=PROBATIONARY_CATEGORIES)
    elif group_filter == 'pending':
        orgs = orgs.filter(status__in=['pending', 'under_review'])
    elif group_filter == 'dissolved':
        orgs = orgs.filter(status='rejected')
    # else: all

    orgs = orgs.order_by('status', 'name')

    return render(request, 'core/admin/orgs.html', {
        'orgs': orgs,
        'group_filter': group_filter,
    })


@admin_required
@ratelimit(key='user', rate='20/h', method='POST', block=True)
def admin_dissolve_org_view(request, org_id):
    from organizations.models import Organization
    org = get_object_or_404(Organization, id=org_id)
    
    # Prevent CSO org deletion
    if org.is_cso:
        messages.error(request, 'Cannot dissolve the CSO organization.')
        return redirect('core:admin_orgs')
    
    if request.method == 'POST':
        # Set status to 'rejected' — Organization.save() will sync is_active=False from this
        org.status = 'rejected'
        org.save()
        
        # Mark all active memberships as removed
        from memberships.models import Membership
        member_count = Membership.objects.filter(organization=org, status='active').count()
        Membership.objects.filter(organization=org, status='active').update(status='removed')
        log_action(
            actor=request.user,
            action=AuditActions.ORG_DISSOLVED,
            target=org,
            details=f'Dissolved organization: {org.name}. Members: {member_count}',
            request=request
        )
        
        messages.success(request, f'"{org.name}" has been dissolved.')
    return redirect('core:admin_orgs')


@admin_required
def admin_categories_view(request):
    from organizations.models import OrganizationCategory
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            if name:
                OrganizationCategory.objects.get_or_create(
                    name=name,
                    defaults={'description': description}
                )
                messages.success(request, f'Category "{name}" created.')
        elif action == 'toggle':
            cat_id = request.POST.get('category_id')
            cat = get_object_or_404(OrganizationCategory, id=cat_id)
            cat.is_active = not cat.is_active
            cat.save()
            messages.success(request, f'Category "{cat.name}" updated.')
        return redirect('core:admin_categories')

    categories = OrganizationCategory.objects.all().order_by('name')
    return render(request, 'core/admin/categories.html', {'categories': categories})



@login_required
def search_view(request):
    from organizations.models import Organization
    from accounts.models import User

    query = request.GET.get('q', '').strip()
    orgs = []
    students = []

    if query:
        orgs = Organization.objects.filter(
            is_active=True,
            status__in=['probationary', 'institutional', 'active', 'renewal_due'],
            name__icontains=query
        ).prefetch_related('categories')[:8]

        students = User.objects.filter(
            is_active=True
        ).filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(student_id__icontains=query)
        )[:8]

    return render(request, 'core/search.html', {
        'query': query,
        'orgs': orgs,
        'students': students,
    })


# ─── CSO member search (for admin privilege grant autocomplete) ───────────────

@login_required
def cso_member_search_view(request):
    """Return JSON list of CSO org members eligible for admin grant, matching name or ID."""
    from django.http import JsonResponse
    from django.db.models import Q
    from accounts.models import User
    from organizations.models import Organization
    from memberships.models import Membership

    if not (request.user.is_cso_president or request.user.is_cso_admin):
        return JsonResponse({'results': []}, status=403)

    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    cso_org = Organization.objects.filter(is_cso=True).first()
    cso_member_ids = []
    if cso_org:
        cso_member_ids = Membership.objects.filter(
            organization=cso_org, status='active'
        ).values_list('user_id', flat=True)

    users = User.objects.filter(
        is_active=True,
        is_cso_admin=False,
        is_cso_president=False,
        id__in=cso_member_ids,
    ).filter(
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q) |
        Q(student_id__icontains=q)
    )[:10]

    results = [
        {'id': u.id, 'name': u.get_full_name(), 'student_id': u.student_id}
        for u in users
    ]
    return JsonResponse({'results': results})


# ─── CSO member search for presidency transfer (returns membership IDs) ───────

@login_required
def cso_successor_search_view(request):
    """Return JSON list of eligible presidency successors (membership ID + user info)."""
    from django.http import JsonResponse
    from django.db.models import Q
    from django.utils import timezone
    from organizations.models import Organization
    from memberships.models import Membership

    if not request.user.is_cso_president:
        return JsonResponse({'results': []}, status=403)

    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    cso_org = Organization.objects.filter(is_cso=True).first()
    if not cso_org:
        return JsonResponse({'results': []})

    memberships = Membership.objects.filter(
        organization=cso_org,
        status='active',
    ).exclude(
        user=request.user,
    ).exclude(
        user__cso_admin_expiry__isnull=False,
    ).filter(
        Q(user__first_name__icontains=q) |
        Q(user__last_name__icontains=q) |
        Q(user__student_id__icontains=q)
    ).select_related('user', 'custom_role')[:10]

    results = [
        {
            'membership_id': m.id,
            'name': m.user.get_full_name(),
            'student_id': m.user.student_id,
            'role': m.custom_role.name if m.custom_role else m.get_role_display(),
        }
        for m in memberships
    ]
    return JsonResponse({'results': results})


# ─── President-only decorator ─────────────────────────────────────────────────

def president_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        if not request.user.is_cso_president:
            messages.error(request, 'This section is restricted to the CSO President.')
            return redirect('core:admin_panel')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─── Admin privilege management ───────────────────────────────────────────────

def president_or_co_chairman_required(view_func):
    """Allow access to CSO President or active CSO co-chairman."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        # Enforce expiry before checking role
        from core.enforcement import expire_cso_admin_for_user
        expire_cso_admin_for_user(request.user)
        if request.user.is_cso_president:
            return view_func(request, *args, **kwargs)
        # Check if they're an active co-chairman in the CSO org
        from organizations.models import Organization
        from memberships.models import Membership
        cso_org = Organization.objects.filter(is_cso=True).first()
        is_co_chairman = cso_org and Membership.objects.filter(
            user=request.user, organization=cso_org,
            role='co_chairman', status='active'
        ).exists()
        if is_co_chairman:
            return view_func(request, *args, **kwargs)
        messages.error(request, 'This section is restricted to the CSO President or Vice Chairman.')
        return redirect('core:admin_panel')
    return wrapper


@president_or_co_chairman_required
@ratelimit(key='user', rate='20/h', method='POST', block=True)
def admin_privileges_view(request):
    from accounts.models import User
    from django.utils import timezone
    from core.enforcement import expire_all_cso_admins

    # Enforce all expired temp admin access
    expire_all_cso_admins()

    admins = User.objects.filter(
        is_active=True
    ).filter(
        models.Q(is_cso_admin=True) | models.Q(is_cso_president=True)
    ).order_by('last_name')

    # Only CSO org members can be granted admin access
    from organizations.models import Organization
    from memberships.models import Membership
    cso_org = Organization.objects.filter(is_cso=True).first()
    cso_member_ids = []
    if cso_org:
        cso_member_ids = Membership.objects.filter(
            organization=cso_org, status='active'
        ).values_list('user_id', flat=True)

    students = User.objects.filter(
        is_active=True, is_cso_admin=False, is_cso_president=False,
        id__in=cso_member_ids,
    ).order_by('last_name')

    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')
        target = get_object_or_404(User, id=user_id, is_active=True)

        if action == 'grant':
            # Server-side guard: block if already admin/president
            if target.is_cso_admin or target.is_cso_president:
                messages.error(request, f'{target.get_full_name()} already has admin access.')
                return redirect('core:admin_privileges')

            # Server-side guard: must be an active CSO org member
            from organizations.models import Organization
            from memberships.models import Membership
            cso_org = Organization.objects.filter(is_cso=True).first()
            is_cso_member = cso_org and Membership.objects.filter(
                user=target, organization=cso_org, status='active'
            ).exists()
            if not is_cso_member:
                messages.error(request, f'{target.get_full_name()} is not an active CSO org member and cannot be granted admin access.')
                return redirect('core:admin_privileges')

            target.is_cso_admin = True
            target.is_manually_granted_admin = True
            target.cso_admin_expiry = None
            target.save(update_fields=['is_cso_admin', 'is_manually_granted_admin', 'cso_admin_expiry'])
            
            # Audit log
            log_action(
                actor=request.user,
                action=AuditActions.ADMIN_GRANTED,
                target=target,
                details=f'Granted CSO admin privileges to {target.get_full_name()} ({target.student_id})',
                request=request
            )
            
            # Auto-add to CSO org as officer (CSO org members are all officers/leaders)
            from organizations.models import Organization
            from memberships.models import Membership
            cso_org = Organization.objects.filter(is_cso=True).first()
            if cso_org:
                Membership.objects.get_or_create(
                    user=target,
                    organization=cso_org,
                    defaults={'role': 'officer', 'status': 'active', 'has_chairman_privileges': False}
                )
            from announcements.utils import send_notification
            from django.urls import reverse
            send_notification(
                title='You have been granted CSO Admin access',
                message=f'{request.user.get_full_name()} has granted you CSO Admin privileges.',
                recipients=[target],
                sender=request.user,
                is_priority=True,
                link_url=reverse('core:admin_panel'),
                notification_type='system',
            )
            messages.success(request, f'{target.get_full_name()} is now a CSO Admin.')

        elif action == 'revoke':
            from organizations.models import Organization
            from memberships.models import Membership
            cso_org = Organization.objects.filter(is_cso=True).first()
            # Always block revoking a co-chairman — their admin is tied to their role
            is_co_chairman = cso_org and Membership.objects.filter(
                user=target, organization=cso_org,
                role='co_chairman', status='active'
            ).exists()
            if is_co_chairman:
                messages.error(request, f'{target.get_full_name()} is a vice chairman — their admin access is tied to their role. Demote them to remove admin access.')
                return redirect('core:admin_privileges')
            target.is_cso_admin = False
            target.is_manually_granted_admin = False
            target.cso_admin_expiry = None
            target.save(update_fields=['is_cso_admin', 'is_manually_granted_admin', 'cso_admin_expiry'])
            
            # Audit log
            log_action(
                actor=request.user,
                action=AuditActions.ADMIN_REVOKED,
                target=target,
                details=f'Revoked CSO admin privileges from {target.get_full_name()} ({target.student_id})',
                request=request
            )
            
            from announcements.utils import send_notification
            from django.urls import reverse
            send_notification(
                title='Your CSO Admin access has been revoked',
                message=f'{request.user.get_full_name()} has revoked your CSO Admin privileges.',
                recipients=[target],
                sender=request.user,
                is_priority=True,
                link_url=reverse('core:dashboard'),
                notification_type='system',
            )
            messages.success(request, f'{target.get_full_name()}\'s admin access revoked.')

        return redirect('core:admin_privileges')

    return render(request, 'core/admin/privileges.html', {
        'admins': admins,
        'students': students,
    })


# ─── System settings ──────────────────────────────────────────────────────────

@president_required
def system_settings_view(request):
    from core.models import SystemSetting

    settings_keys = ['admin_guide', 'welcome_message', 'system_rules']
    settings = {s.key: s for s in SystemSetting.objects.filter(key__in=settings_keys)}

    if request.method == 'POST':
        for key in settings_keys:
            value = request.POST.get(key, '').strip()
            obj, _ = SystemSetting.objects.get_or_create(key=key, defaults={'value': value})
            obj.value = value
            obj.last_updated_by = request.user
            obj.save()
        messages.success(request, 'System settings updated.')
        return redirect('core:system_settings')

    return render(request, 'core/admin/system_settings.html', {
        'settings': settings,
    })


# ─── Presidency transfer ──────────────────────────────────────────────────────

@president_required
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def presidency_transfer_view(request):
    from accounts.models import User
    from datetime import timedelta
    from django.utils import timezone
    from core.models import HandoverNote

    # Eligible successors — active CSO org members who are not the current president
    # Exclude anyone currently in a 24hr temp admin window (outgoing president)
    from organizations.models import Organization
    from memberships.models import Membership
    cso_org = Organization.objects.filter(is_cso=True).first()
    eligible = []
    if cso_org:
        eligible = Membership.objects.filter(
            organization=cso_org,
            status='active',
        ).exclude(
            user=request.user,
        ).exclude(
            user__cso_admin_expiry__isnull=False,
        ).select_related('user', 'custom_role').order_by('user__last_name')

    if request.method == 'POST':
        from django.db import transaction
        
        successor_id = request.POST.get('successor_id', '').strip()
        note_text = request.POST.get('note', '').strip()

        if not successor_id:
            messages.error(request, 'Please select a successor.')
            return render(request, 'core/admin/presidency_transfer.html', {'eligible': eligible})

        # successor_id is a membership ID from the eligible queryset
        from organizations.models import Organization
        from memberships.models import Membership
        cso_org = Organization.objects.filter(is_cso=True).first()
        try:
            successor_membership = Membership.objects.get(
                id=successor_id, organization=cso_org, status='active'
            )
        except Membership.DoesNotExist:
            messages.error(request, 'Invalid successor selected.')
            return render(request, 'core/admin/presidency_transfer.html', {'eligible': eligible})

        successor = successor_membership.user

        if successor == request.user:
            messages.error(request, 'You cannot transfer the presidency to yourself.')
            return render(request, 'core/admin/presidency_transfer.html', {'eligible': eligible})

        if successor.cso_admin_expiry:
            messages.error(request, 'This person is currently in a 24hr transition window and cannot be selected as successor.')
            return render(request, 'core/admin/presidency_transfer.html', {'eligible': eligible})

        # Use transaction with row locking to prevent race conditions
        with transaction.atomic():
            # Lock current president row to prevent concurrent transfers
            current_president = User.objects.select_for_update().get(id=request.user.id)
            
            # Verify still president after lock acquired
            if not current_president.is_cso_president:
                messages.error(request, 'You are no longer the CSO President.')
                return redirect('core:admin_panel')

            # 1. Grant successor president status + make them chairman of CSO org
            successor.is_cso_president = True
            successor.is_cso_admin = True
            successor.cso_admin_expiry = None
            successor.save(update_fields=['is_cso_president', 'is_cso_admin', 'cso_admin_expiry'])

            cso_org = cso_org  # already fetched above
            if cso_org:
                Membership.objects.update_or_create(
                    user=successor,
                    organization=cso_org,
                    defaults={
                        'role': 'chairman',
                        'status': 'active',
                        'has_chairman_privileges': True,
                        'co_chairman_expiry': None,
                    }
                )

            # 2. Outgoing president → co-chairman in CSO org for 24hrs + temp admin
            request.user.is_cso_president = False
            request.user.is_cso_admin = True
            request.user.cso_admin_expiry = timezone.now() + timedelta(hours=24)
            request.user.save(update_fields=['is_cso_president', 'is_cso_admin', 'cso_admin_expiry'])

            if cso_org:
                outgoing_membership, _ = Membership.objects.update_or_create(
                    user=request.user,
                    organization=cso_org,
                    defaults={
                        'role': 'co_chairman',
                        'status': 'active',
                        'has_chairman_privileges': True,
                        'co_chairman_expiry': timezone.now() + timedelta(hours=24),
                    }
                )

            # 3. Save handover note
            if note_text:
                HandoverNote.objects.create(
                    from_user=request.user,
                    to_user=successor,
                    organization=None,
                    type='cso_president',
                    note=note_text,
                )

        # Audit log (outside transaction)
        log_action(
            actor=request.user,
            action=AuditActions.PRESIDENCY_TRANSFER,
            target=successor,
            details=f'Transferred presidency from {request.user.get_full_name()} to {successor.get_full_name()}. Handover note: {note_text[:100] if note_text else "None"}',
            request=request
        )
        
        log_action(
            actor=request.user,
            action=AuditActions.TEMP_ADMIN_GRANTED,
            target=request.user,
            details=f'Granted 24-hour temporary admin access. Expires: {request.user.cso_admin_expiry}',
            request=request
        )

        # Schedule auto-revoke via Celery (outside transaction)
        try:
            from core.tasks import expire_temp_admin_access
            expire_temp_admin_access.apply_async(
                args=[request.user.id],
                eta=request.user.cso_admin_expiry,
            )
        except Exception:
            pass

        # 4. Notifications
        from announcements.utils import send_notification
        from django.urls import reverse
        msg = f'{request.user.get_full_name()} has transferred the CSO Presidency to you.'
        if note_text:
            msg += f'\n\nHandover note: {note_text}'
        send_notification(
            title='You are now the CSO President',
            message=msg,
            recipients=[successor],
            sender=request.user,
            is_priority=True,
            link_url=reverse('core:admin_panel'),
            notification_type='system',
        )
        send_notification(
            title='Presidency transferred',
            message=(
                f'You have transferred the CSO Presidency to {successor.get_full_name()}. '
                f'You have temporary CSO Admin access for 24 hours.'
            ),
            recipients=[request.user],
            sender=request.user,
            is_priority=True,
            link_url=reverse('core:dashboard'),
            notification_type='system',
        )

        messages.success(
            request,
            f'Presidency transferred to {successor.get_full_name()}. '
            f'You have 24 hours of temporary admin access.'
        )
        return redirect('core:admin_panel')

    return render(request, 'core/admin/presidency_transfer.html', {
        'eligible': eligible,
    })


@login_required
def revoke_temp_admin_view(request):
    """Outgoing president self-revokes their 24hr temp admin access."""
    if not request.user.is_cso_admin or not request.user.cso_admin_expiry:
        messages.error(request, 'You do not have temporary admin access to revoke.')
        return redirect('core:dashboard')

    if request.method == 'POST':
        from organizations.models import Organization
        from memberships.models import Membership

        cso_org = Organization.objects.filter(is_cso=True).first()
        if cso_org:
            m = Membership.objects.filter(
                user=request.user, organization=cso_org,
                role='co_chairman', co_chairman_expiry__isnull=False, status='active'
            ).first()
            if m:
                if m.pending_role:
                    m.role = m.pending_role
                    m.has_chairman_privileges = m.pending_role == 'co_chairman'
                    m.custom_role = m.pending_custom_role
                    m.pending_role = None
                    m.pending_custom_role = None
                else:
                    # CSO outgoing president becomes officer, not member
                    m.role = 'officer'
                    m.has_chairman_privileges = False
                    m.custom_role = None
                m.co_chairman_expiry = None
                m.save()

        request.user.is_cso_admin = False
        request.user.is_manually_granted_admin = False
        request.user.cso_admin_expiry = None
        request.user.save(update_fields=['is_cso_admin', 'is_manually_granted_admin', 'cso_admin_expiry'])
        messages.success(request, 'Your temporary admin access has been revoked.')
        return redirect('core:dashboard')

    # GET — show a dedicated confirmation page
    return render(request, 'core/revoke_temp_admin_confirm.html')


# ─── Handover notes history ───────────────────────────────────────────────────

@president_required
def handover_notes_view(request):
    from core.models import HandoverNote
    notes = HandoverNote.objects.all().select_related(
        'from_user', 'to_user', 'organization'
    ).order_by('-created_at')
    return render(request, 'core/admin/handover_notes.html', {'notes': notes})


# ─── Profile correction requests ─────────────────────────────────────────────

LOCKED_FIELD_LABELS = {
    'first_name': 'First name',
    'middle_initial': 'Middle initial',
    'last_name': 'Last name',
    'student_id': 'Student ID',
    'course': 'Course',
    'year_level': 'Year level',
}


@admin_required
def admin_correction_requests_view(request):
    from accounts.models import ProfileCorrectionRequest
    all_requests = ProfileCorrectionRequest.objects.select_related(
        'user', 'user__course', 'reviewed_by'
    ).order_by(
        models.Case(
            models.When(status='pending', then=0),
            default=1,
            output_field=models.IntegerField(),
        ),
        '-created_at',
    )
    pending_count = all_requests.filter(status='pending').count()
    return render(request, 'core/admin/correction_requests.html', {
        'correction_requests': all_requests,
        'pending_count': pending_count,
        'locked_field_labels': LOCKED_FIELD_LABELS,
    })


@admin_required
@ratelimit(key='user', rate='20/h', method='POST', block=True)
def admin_review_correction_request_view(request, request_id):
    from accounts.models import ProfileCorrectionRequest, Course
    from announcements.utils import send_notification
    from django.utils import timezone

    correction_request = get_object_or_404(ProfileCorrectionRequest, id=request_id)

    if request.method != 'POST':
        return redirect('core:admin_correction_requests')

    if correction_request.status != 'pending':
        messages.error(request, 'This request has already been reviewed.')
        return redirect('core:admin_correction_requests')

    action = request.POST.get('action')

    if action == 'approve':
        student = correction_request.user
        field = correction_request.field_name
        new_value = correction_request.new_value

        # Apply the field update
        if field == 'course':
            try:
                course = Course.objects.get(id=int(new_value))
                setattr(student, field, course)
            except (Course.DoesNotExist, ValueError):
                messages.error(request, 'Could not resolve the course. Please check the new value.')
                return redirect('core:admin_correction_requests')
        elif field == 'year_level':
            try:
                setattr(student, field, int(new_value))
            except ValueError:
                messages.error(request, 'Invalid year level value.')
                return redirect('core:admin_correction_requests')
        elif field == 'middle_initial':
            # Strip any trailing dot and uppercase before saving
            setattr(student, field, new_value.strip().rstrip('.').upper())
        else:
            setattr(student, field, new_value)
        student.save()

        # Update the request
        correction_request.status = 'approved'
        correction_request.reviewed_by = request.user
        correction_request.reviewed_at = timezone.now()
        correction_request.save()
        
        # Audit log
        log_action(
            actor=request.user,
            action=AuditActions.CORRECTION_APPROVED,
            target=student,
            details=f'Approved correction for {student.get_full_name()}: {field} changed from "{correction_request.old_value}" to "{new_value}"',
            request=request
        )

        # Supersede other pending requests for the same field+user
        ProfileCorrectionRequest.objects.filter(
            user=student,
            field_name=field,
            status='pending',
        ).exclude(id=correction_request.id).update(
            status='rejected',
            rejection_reason='Superseded by an approved correction request for the same field.',
            reviewed_by=request.user,
            reviewed_at=timezone.now(),
        )

        # Notify student
        try:
            label = LOCKED_FIELD_LABELS.get(field, field)
            from django.urls import reverse
            send_notification(
                title='Correction request approved',
                message=f'Your correction request for {label} has been approved. Your profile has been updated.',
                recipients=[student],
                sender=request.user,
                is_priority=True,
                link_url=reverse('accounts:profile'),
                notification_type='profile',
            )
        except Exception:
            pass

        messages.success(request, f'Correction request approved and profile updated.')

    elif action == 'reject':
        rejection_reason = request.POST.get('rejection_reason', '').strip()
        if not rejection_reason:
            messages.error(request, 'A rejection reason is required.')
            return redirect('core:admin_correction_requests')

        correction_request.status = 'rejected'
        correction_request.reviewed_by = request.user
        correction_request.reviewed_at = timezone.now()
        correction_request.rejection_reason = rejection_reason
        correction_request.save()
        
        # Audit log
        log_action(
            actor=request.user,
            action=AuditActions.CORRECTION_REJECTED,
            target=correction_request.user,
            details=f'Rejected correction for {correction_request.user.get_full_name()}: {correction_request.field_name}. Reason: {rejection_reason}',
            request=request
        )

        # Notify student
        try:
            label = LOCKED_FIELD_LABELS.get(correction_request.field_name, correction_request.field_name)
            from django.urls import reverse
            send_notification(
                title='Correction request rejected',
                message=f'Your correction request for {label} was rejected. Reason: {rejection_reason}',
                recipients=[correction_request.user],
                sender=request.user,
                is_priority=True,
                link_url=reverse('accounts:profile'),
                notification_type='profile',
            )
        except Exception:
            pass

        messages.success(request, 'Correction request rejected.')

    return redirect('core:admin_correction_requests')


# ─── Academic periods ─────────────────────────────────────────────────────────

@admin_required
def admin_academic_periods_view(request):
    from core.models import AcademicPeriod

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            name = request.POST.get('name', '').strip()
            start_date = request.POST.get('start_date', '').strip()
            end_date = request.POST.get('end_date', '').strip()
            is_current = request.POST.get('is_current') == 'on'

            if not name or not start_date or not end_date:
                messages.error(request, 'Name, start date, and end date are required.')
            else:
                AcademicPeriod.objects.create(
                    name=name,
                    start_date=start_date,
                    end_date=end_date,
                    is_current=is_current,
                    created_by=request.user,
                )
                messages.success(request, f'Academic period "{name}" created.')

        elif action == 'set_current':
            period_id = request.POST.get('period_id')
            period = get_object_or_404(AcademicPeriod, id=period_id)
            AcademicPeriod.objects.filter(is_current=True).update(is_current=False)
            period.is_current = True
            period.save(update_fields=['is_current'])
            messages.success(request, f'"{period.name}" is now the current academic period.')

        elif action == 'delete':
            period_id = request.POST.get('period_id')
            period = get_object_or_404(AcademicPeriod, id=period_id)
            if period.is_current:
                messages.error(request, 'Cannot delete the current academic period.')
            else:
                name = period.name
                period.delete()
                messages.success(request, f'"{name}" deleted.')

        return redirect('core:admin_academic_periods')

    periods = AcademicPeriod.objects.order_by('-start_date')
    return render(request, 'core/admin/academic_periods.html', {'periods': periods})


@admin_required
@ratelimit(key='user', rate='5/d', method='POST', block=True)
def admin_school_year_transition_view(request):
    """
    Trigger school year transition (unified):
    - Increment all active student year levels by 1
    - Flag students at year 5+ for admin review (possible graduates)
    - Demote all org officers to members (excluding CSO org)
    - Flip all active/probationary/institutional orgs to renewal_due
    - Set optional renewal deadline in SystemSetting
    - Notify all affected org chairmen
    """
    from accounts.models import User
    from memberships.models import Membership
    from django.utils import timezone
    from organizations.models import Organization

    if request.method == 'POST':
        confirm = request.POST.get('confirm') == 'yes'
        if not confirm:
            messages.error(request, 'Please confirm the transition.')
            return redirect('core:admin_school_year_transition')

        graduate_threshold = int(request.POST.get('graduate_threshold', 4))
        renewal_deadline = request.POST.get('renewal_deadline', '').strip()

        # 1. Bulk-increment year levels for all active students
        from django.db.models import F
        incremented = User.objects.filter(is_active=True).update(year_level=F('year_level') + 1)
        flagged = User.objects.filter(is_active=True, year_level__gt=graduate_threshold).count()

        # 2. Demote all officers to members across all orgs, excluding the CSO org
        cso_org = Organization.objects.filter(is_cso=True).first()
        demotion_qs = Membership.objects.filter(role='officer', status='active')
        if cso_org:
            demotion_qs = demotion_qs.exclude(organization=cso_org)
        demoted = demotion_qs.update(role='member', has_chairman_privileges=False, custom_role=None)

        # Clear adviser_since for all non-adviser memberships so the new period starts fresh.
        # Advisers who are still active advisers keep their adviser_since until demoted.
        Membership.objects.filter(
            status='active'
        ).exclude(role='adviser').update(adviser_since=None)

        # 3. Store renewal deadline in SystemSetting and set cycle flags
        from core.models import SystemSetting
        if renewal_deadline:
            obj, _ = SystemSetting.objects.get_or_create(
                key='renewal_deadline', defaults={'value': renewal_deadline}
            )
            obj.value = renewal_deadline
            obj.last_updated_by = request.user
            obj.save()

        # Mark that a school year transition happened, reset the "closed" flag
        SystemSetting.objects.update_or_create(
            key='school_year_transitioned',
            defaults={'value': 'true', 'last_updated_by': request.user},
        )
        # Clear the renewal_period_closed flag so Close Renewal can be used again this cycle
        SystemSetting.objects.filter(key='renewal_period_closed').delete()

        # 4. Flip eligible orgs to renewal_due and notify chairmen
        eligible_statuses = ['probationary', 'institutional', 'active']
        eligible_orgs = Organization.objects.filter(status__in=eligible_statuses, is_cso=False)
        org_count = eligible_orgs.count()

        try:
            from announcements.utils import send_notification
            from django.urls import reverse as _reverse
            for org in eligible_orgs:
                org.pre_renewal_status = org.status
                org.status = 'renewal_due'
                org.save()

                chairman_membership = Membership.objects.filter(
                    organization=org, role='chairman', status='active'
                ).select_related('user').first()
                if chairman_membership:
                    deadline_str = f' Renewal deadline: {renewal_deadline}.' if renewal_deadline else ''
                    send_notification(
                        title=f'Renewal required — {org.name}',
                        message=(
                            f'A new school year has begun. Your organization "{org.name}" must submit '
                            f'a renewal application to maintain its accredited status.{deadline_str}'
                        ),
                        recipients=[chairman_membership.user],
                        sender=request.user,
                        organization=org,
                        is_priority=True,
                        link_url=_reverse('organizations:renewal_apply', kwargs={'org_id': org.id}),
                        notification_type='organization',
                    )
        except Exception:
            pass

        # 5. Audit log
        log_action(
            actor=request.user,
            action=AuditActions.YEAR_TRANSITION,
            details=(
                f'School year transition. Students incremented: {incremented}, '
                f'Graduates flagged: {flagged}, Officers demoted: {demoted}, '
                f'Orgs set to renewal_due: {org_count}. '
                f'Deadline: {renewal_deadline or "not set"}.'
            ),
            request=request,
        )

        # 6. Notify all students
        try:
            from announcements.utils import send_notification
            from django.urls import reverse as _reverse
            all_students = list(User.objects.filter(is_active=True))
            if all_students:
                send_notification(
                    title='School year transition completed',
                    message='A new school year has begun. Your year level has been updated.',
                    recipients=all_students,
                    sender=request.user,
                    is_priority=True,
                    link_url=_reverse('accounts:profile'),
                    notification_type='system',
                )
        except Exception:
            pass

        messages.success(
            request,
            f'School year transition complete. '
            f'{incremented} student year levels incremented. '
            f'{flagged} possible graduate(s) flagged (year level > {graduate_threshold}). '
            f'{demoted} officer(s) demoted to member. '
            f'{org_count} org(s) set to renewal required.'
            + (f' Deadline: {renewal_deadline}.' if renewal_deadline else '')
        )
        return redirect('core:admin_panel')

    # GET — show confirmation page with impact summary
    student_count = User.objects.filter(is_active=True).count()
    cso_org = Organization.objects.filter(is_cso=True).first()
    officer_count_qs = Membership.objects.filter(role='officer', status='active')
    if cso_org:
        officer_count_qs = officer_count_qs.exclude(organization=cso_org)
    officer_count = officer_count_qs.count()
    eligible_org_count = Organization.objects.filter(
        status__in=['probationary', 'institutional', 'active'], is_cso=False
    ).count()
    return render(request, 'core/admin/school_year_transition.html', {
        'student_count': student_count,
        'officer_count': officer_count,
        'eligible_org_count': eligible_org_count,
    })


# ─── Discovery widget (HTMX partial) ─────────────────────────────────────────

@login_required
def discovery_widget_view(request):
    tab = request.GET.get('tab', 'padlet')
    context = {'tab': tab}

    def get_board_gradient(board):
        """Return the CSS gradient string for a board based on scope and board_color."""
        if board.scope == 'system':
            return 'linear-gradient(to bottom right, #f97316, #ea580c, #dc2626)'
        # Org board — use board_color
        color_map = {
            'blue-slate': 'linear-gradient(to bottom right, #6a8fc5, #4a6fa5, #2d4a7a)',
            'blue-ocean': 'linear-gradient(to bottom right, #3ea8e8, #1e88c8, #0d5a8a)',
            'blue-denim': 'linear-gradient(to bottom right, #5a7faa, #3a5f8a, #1e3a5f)',
            'blue-sky': 'linear-gradient(to bottom right, #7bd8f8, #5bb8e8, #2a8abf)',
            'green-sage': 'linear-gradient(to bottom right, #9acbaa, #7aab8a, #4a7a5a)',
            'green-forest': 'linear-gradient(to bottom right, #5a9a6a, #3a7a4a, #1e4a2a)',
            'green-mint': 'linear-gradient(to bottom right, #8eefca, #6ecfa8, #3aaa7a)',
            'green-olive': 'linear-gradient(to bottom right, #aabc6a, #8a9a4a, #5a6a2a)',
            'purple-grape': 'linear-gradient(to bottom right, #9a6aba, #7a4a9a, #4a2a6a)',
            'purple-lavender': 'linear-gradient(to bottom right, #c09ae8, #a07ac8, #6a4a9a)',
            'purple-plum': 'linear-gradient(to bottom right, #8a5a9a, #6a3a7a, #3a1a4a)',
            'red-rose': 'linear-gradient(to bottom right, #e87a9a, #c85a7a, #8a2a4a)',
            'red-crimson': 'linear-gradient(to bottom right, #e0594b, #c0392b, #7a1a10)',
            'pink-blush': 'linear-gradient(to bottom right, #f8c0d8, #e8a0b8, #c06080)',
            'warm-sand': 'linear-gradient(to bottom right, #e8c89a, #c8a87a, #8a6a3a)',
            'warm-terracotta': 'linear-gradient(to bottom right, #e89a7a, #c87a5a, #8a4a2a)',
            'warm-caramel': 'linear-gradient(to bottom right, #e8ba7a, #c89a5a, #8a5a1a)',
            'dark-charcoal': 'linear-gradient(to bottom right, #6a6a7a, #4a4a5a, #1a1a2a)',
            'dark-midnight': 'linear-gradient(to bottom right, #2a3a5a, #1a2a4a, #0a0a1a)',
        }
        return color_map.get(board.board_color, 'linear-gradient(to bottom right, #3B5BBE, #1B2A7B, #0d1a4a)')

    def get_board_accent(board):
        """Return a single mid-tone hex color for badge/button accents, matching the board gradient."""
        if board.scope == 'system':
            return '#ea580c'  # orange-600
        accent_map = {
            'blue-slate': '#4a6fa5',
            'blue-ocean': '#1e88c8',
            'blue-denim': '#3a5f8a',
            'blue-sky': '#5bb8e8',
            'green-sage': '#7aab8a',
            'green-forest': '#3a7a4a',
            'green-mint': '#6ecfa8',
            'green-olive': '#8a9a4a',
            'purple-grape': '#7a4a9a',
            'purple-lavender': '#a07ac8',
            'purple-plum': '#6a3a7a',
            'red-rose': '#c85a7a',
            'red-crimson': '#c0392b',
            'pink-blush': '#e8a0b8',
            'warm-sand': '#c8a87a',
            'warm-terracotta': '#c87a5a',
            'warm-caramel': '#c89a5a',
            'dark-charcoal': '#4a4a5a',
            'dark-midnight': '#1a2a4a',
        }
        return accent_map.get(board.board_color, '#3B5BBE')

    if tab == 'padlet':
        from padlet.models import Board
        from memberships.models import Membership
        from django.db.models import Count, Q

        # Get user's organization IDs
        user_org_ids = list(Membership.objects.filter(
            user=request.user, status='active', organization__is_active=True,
        ).values_list('organization_id', flat=True))

        # Get 1 system-wide board with most notes
        system_board = Board.objects.filter(
            scope='system',
            status='active',
        ).annotate(
            note_count=Count('posts'),
            contributor_count=Count('posts__author', distinct=True)
        ).select_related('organization', 'created_by').prefetch_related(
            'posts__author'
        ).order_by('-note_count').first()

        # Get 1 org board relevant to this user:
        # - boards from orgs they're a member of, OR
        # - boards they created, OR
        # - all org boards if CSO admin/president
        is_cso_admin = request.user.is_cso_admin or request.user.is_cso_president
        if is_cso_admin:
            org_board_qs = Board.objects.filter(scope='org', status='active')
        else:
            org_board_qs = Board.objects.filter(
                Q(scope='org', organization_id__in=user_org_ids) |
                Q(scope='org', created_by=request.user),
                status='active',
            )

        org_board = org_board_qs.annotate(
            note_count=Count('posts'),
            contributor_count=Count('posts__author', distinct=True)
        ).select_related('organization', 'created_by').prefetch_related(
            'posts__author'
        ).order_by('-note_count').first()

        # Build boards list: system board first, then org board
        boards = []
        if system_board:
            top_contributors = system_board.posts.values('author').annotate(
                post_count=Count('id')
            ).order_by('-post_count')[:3]
            contributor_ids = [c['author'] for c in top_contributors if c['author']]
            from accounts.models import User
            system_board.top_contributors = list(User.objects.filter(id__in=contributor_ids))
            system_board.user_already_posted = system_board.posts.filter(author=request.user).exists()
            system_board.user_is_manager = request.user.is_cso_admin or request.user.is_cso_president
            system_board.board_gradient = get_board_gradient(system_board)
            system_board.board_accent = get_board_accent(system_board)
            boards.append(system_board)

        if org_board:
            top_contributors = org_board.posts.values('author').annotate(
                post_count=Count('id')
            ).order_by('-post_count')[:3]
            contributor_ids = [c['author'] for c in top_contributors if c['author']]
            from accounts.models import User
            org_board.top_contributors = list(User.objects.filter(id__in=contributor_ids))
            org_board.user_already_posted = org_board.posts.filter(author=request.user).exists()
            # Manager = board creator OR org chairman/co-chairman
            from memberships.models import Membership as _Mbr
            org_board.user_is_manager = (
                org_board.created_by == request.user or
                _Mbr.objects.filter(
                    user=request.user,
                    organization=org_board.organization,
                    role__in=['chairman', 'co_chairman'],
                    status='active',
                    organization__is_active=True,
                ).exists()
            )
            org_board.board_gradient = get_board_gradient(org_board)
            org_board.board_accent = get_board_accent(org_board)
            boards.append(org_board)

        context['boards'] = boards
        context['user_org_ids'] = user_org_ids

    elif tab == 'elections':
        from elections.models import Election, ElectionVoter, Vote, Candidate
        from django.db.models import Count
        open_elections = Election.objects.filter(
            voters__user=request.user,
            status='open',
        ).select_related('organization').order_by('-created_at')[:5]

        voted_election_ids = set(
            Vote.objects.filter(voter=request.user).values_list('election_id', flat=True)
        )

        # For each election, get positions with candidates and vote counts
        elections_data = []
        for election in open_elections:
            total_voters = ElectionVoter.objects.filter(election=election).count()
            total_votes_cast = Vote.objects.filter(election=election).values('voter').distinct().count()
            positions = election.positions.prefetch_related(
                models.Prefetch(
                    'candidates',
                    queryset=Candidate.objects.select_related('user').annotate(
                        vote_count=Count('votes')
                    ).order_by('-vote_count')
                )
            ).all()
            elections_data.append({
                'election': election,
                'total_voters': total_voters,
                'total_votes_cast': total_votes_cast,
                'positions': positions,
                'has_voted': election.id in voted_election_ids,
                'is_live': True,
            })

        # If no open elections, show recent results_released elections the user was a voter in
        recent_results = []
        if not elections_data:
            recent_elections = Election.objects.filter(
                voters__user=request.user,
                status='results_released',
            ).select_related('organization').order_by('-results_released_at')[:3]

            from elections.utils import get_vote_tally
            for election in recent_elections:
                tally = get_vote_tally(election)
                positions_summary = []
                for position in election.positions.all():
                    pos_data = tally.get(str(position.id), {})
                    candidates = pos_data.get('candidates', [])
                    is_tied = pos_data.get('is_tied', False)
                    no_contest = pos_data.get('no_contest', False)
                    positions_summary.append({
                        'name': position.name,
                        'winner': candidates[0] if candidates and not is_tied and not no_contest else None,
                        'is_tied': is_tied,
                        'no_contest': no_contest,
                    })
                recent_results.append({
                    'election': election,
                    'positions_summary': positions_summary,
                    'is_live': False,
                })

        context['elections_data'] = elections_data
        context['recent_results'] = recent_results
        context['voted_election_ids'] = voted_election_ids

    elif tab == 'leaderboard':
        from reports.models import AccomplishmentReport
        from organizations.models import Organization
        from django.db.models import Count
        orgs = Organization.objects.filter(
            is_active=True, is_cso=False
        ).annotate(
            report_count=Count('accomplishment_reports', filter=models.Q(accomplishment_reports__status='approved'))
        ).order_by('-report_count', 'name')[:5]
        ranked = [{'rank': i + 1, 'org': org, 'report_count': org.report_count} for i, org in enumerate(orgs)]
        context['ranked'] = ranked

    return render(request, 'core/_discovery_widget.html', context)


# ─── Audit log ────────────────────────────────────────────────────────────────

@admin_required
def admin_audit_log_view(request):
    from core.models import AuditLog
    logs = AuditLog.objects.select_related('actor').order_by('-created_at')[:200]
    return render(request, 'core/admin/audit_log.html', {'logs': logs})


# ─── Data export ─────────────────────────────────────────────────────────────

@admin_required
def admin_export_view(request):
    """Export landing page — choose what to export and format."""
    return render(request, 'core/admin/export.html')


@admin_required
def export_students_view(request):
    """Export all active students as CSV or Excel."""
    fmt = request.GET.get('format', 'csv')
    from accounts.models import User
    students = User.objects.filter(is_active=True).select_related('course').order_by('last_name', 'first_name')

    headers = ['Student ID', 'Last Name', 'First Name', 'Middle Initial', 'Email',
               'Course', 'Year Level', 'Contact Number', 'Date Joined']

    rows = []
    for s in students:
        rows.append([
            s.student_id,
            s.last_name,
            s.first_name,
            s.middle_initial or '',
            s.email,
            s.course.abbreviation if s.course else '',
            s.year_level,
            s.contact_number or '',
            s.date_joined.strftime('%Y-%m-%d') if s.date_joined else '',
        ])

    if fmt == 'excel':
        return _excel_response('students', headers, rows)
    return _csv_response('students', headers, rows)


@admin_required
def export_orgs_view(request):
    """Export all active organizations as CSV or Excel."""
    fmt = request.GET.get('format', 'csv')
    from organizations.models import Organization
    from memberships.models import Membership
    from django.db.models import Count

    orgs = Organization.objects.filter(is_active=True, is_cso=False).prefetch_related('categories').annotate(
        member_count=Count('memberships', filter=models.Q(memberships__status='active', memberships__user__is_active=True)),
        approved_reports=Count('accomplishment_reports', filter=models.Q(accomplishment_reports__status='approved')),
    ).order_by('name')

    headers = ['Organization Name', 'Categories', 'Member Count', 'Approved Reports',
               'Founded Year', 'Date Created']

    rows = []
    for org in orgs:
        cats = ', '.join(c.name for c in org.categories.all())
        rows.append([
            org.name,
            cats,
            org.member_count,
            org.approved_reports,
            org.founded_year or '',
            org.date_created.strftime('%Y-%m-%d') if org.date_created else '',
        ])

    if fmt == 'excel':
        return _excel_response('organizations', headers, rows)
    return _csv_response('organizations', headers, rows)


def _csv_response(name, headers, rows):
    import csv
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{name}_export.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def _excel_response(name, headers, rows):
    """Generate an Excel file using openpyxl if available, fallback to CSV."""
    from django.http import HttpResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = name.capitalize()

        # Header row styling
        header_fill = PatternFill(start_color='1B2A7B', end_color='1B2A7B', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{name}_export.xlsx"'
        wb.save(response)
        return response

    except ImportError:
        # openpyxl not installed — fall back to CSV with a note
        return _csv_response(name, headers, rows)
